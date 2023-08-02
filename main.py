import json
import os
import openpyxl

import requests

headers = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "ru,en-US;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
    "Cookie": "s=6c034a0ac6fb42c4805567f284f45c6d; _gat=1; __sreff=1690956007430.1690956007430.1; __reff=[[www.workle.ru/id718491/promopage/realty/]]kwork.ru&1690956007430.1690956007430.1; _ga_1CW549PSHR=GS1.1.1690956007.1.0.1690956007.60.0.0; _ga=GA1.1.7ef66485-b6d4-4d6a-a9f7-05655a612d37; kvcd=1690956007874; km_ai=APLtBdl5ijnj0TAuV8psxpe%2BN%2FA%3D; km_vs=1; km_lv=1690956008; tmr_lvid=84f733767d634ea8fab67d967b2a12bb; tmr_lvidTS=1690956007914; tmr_detect=1%7C1690956007938; d5e8eaeea28b343a46f9b78f17212f99_hits=1; d5e8eaeea28b343a46f9b78f17212f99_vc=1; _gcl_au=1.1.561397774.1690956008",
    "Host": "www.workle.ru",
    "Referer": "https://www.workle.ru/id718491/promopage/realty/?spid=5444",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
    "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1"
}


def getFirstHouseIds(regionId, take, spid=5444):  # 78, 266, 5444
    print(f"[*] Пытаюсь получить ID всех ЖК в городе #{regionId}. ")
    url = f"https://www.workle.ru/api/v1/realty/offers/?regionId={regionId}&skip=0&take={take}&spid={spid}"
    r = requests.get(url=url, headers=headers)
    print(f"[*] Ответ на запрос: {r.status_code}.")
    print(f"[/] Выявляю ID конкретных зданий...")
    jString = json.loads(r.text)
    houseIds = list()
    for h in range(len(jString['list'])):
        houseIds.append(jString['list'][h]['buildingId'])
    print(f"[$] ID зданий классифицированы успешно!")

    return houseIds


def getCurrentHouseInfo(id, spid=5444):
    print(f"[*] Делаю запрос к строению #{id}..")
    url = f"https://www.workle.ru/api/v1/realty/buildingoffers/?buildingId={id}&spid={spid}"
    r = requests.get(url=url, headers=headers)
    print(f"[*] Ответ на запрос: {r.status_code}.")
    print(f"[/] Стуктурирую информацию..")
    jString = json.loads(r.text)
    print(jString)
    data = {
        "BuildingId": id,
        "MainImages": jString[0]['images'],
        "House": "ЖК «" + jString[0]['complexName'] + "»",
        "MinPrice": jString[0]['minPrice'],
        "MaxPrice": jString[0]['maxPrice'],
        "Metro": jString[0]['metro'],
        "MetroTime": jString[0]['metroTime'],
        "Region": jString[0]['region'],
        "Zastroyshik": jString[0]['builderName'],
        "Sdacha": jString[0]['builtDate'],
        "FloorsTotal": jString[0]['floorsTotal'],
        "Otdelka": jString[0]['renovation'],
        "HouseType": jString[0]['buildingType'],
        "Description": jString[0]['description'].split('\n')[1],
        "MinArea": jString[0]['minArea'],
        "MaxArea": jString[0]['maxArea'],
        "SubBuildings": getSubBuildings(id)
    }

    print("[*] Перехожу к загрузке лицевых картинок..")

    mainImgPaths = list()

    for h in range(len(data['MainImages'])):
        url = data['MainImages'][h]['value']
        dlImage(url, str(h), id, "main")
        path = f"images/{id}/{h}.jpg"
        mainImgPaths.append(path)

    print("[/] Лицевые картинки скачаны успешно!")

    data['MainImages'] = ";".join(mainImgPaths)

    print("[*] Перехожу к загрузке плановых картинок..")

    subImgPaths = list()

    for h in range(len(data['SubBuildings'])):
        url = data['SubBuildings'][h]['image']
        dlImage(url, h, id, "sub")
        path = f"images/{id}/subBuilding/{h}.jpg"
        subImgPaths.append(path)

    print("[/] Плановые картинки скачаны успешно!")

    data['SubBuildings'] = ";".join(subImgPaths)

    print(f"[$] Объявление классифицированно успешно! ----------------")
    return data


def dlImage(url, name, buildingId, type="main"):
    print(f"[*] Скачиваю изображение #{buildingId} - {name}.jpg...")

    if not os.path.exists(f"images"):
        os.mkdir("images")

    if type == "main":
        if not os.path.exists(f"images/{buildingId}"):
            os.mkdir(f"images/{buildingId}")

        try:
            img_data = requests.get(url).content
            with open(f'images/{buildingId}/{name}.jpg', 'wb') as handler:
                handler.write(img_data)

            print(f"[$] Фотография скачана успешно!")
        except Exception as e:
            print(f"[!] Произошла ошибка при скачивании фото по ссылке: {url}! [{e}]")

    if type == "sub":
        if not os.path.exists(f"images/{buildingId}"):
            os.mkdir(f"images/{buildingId}")

        if not os.path.exists(f"images/{buildingId}/subBuilding"):
            os.mkdir(f"images/{buildingId}/subBuilding")

        try:
            img_data = requests.get(url).content
            with open(f'images/{buildingId}/subBuilding/{name}.jpg', 'wb') as handler:
                handler.write(img_data)
            print(f"[$] Фотография скачана успешно!")
        except Exception as e:
            print(f"[!] Произошла ошибка при скачивании фото по ссылке: {url}! [{e}]")


def getSubBuildings(buildingId, spid=5444):
    print(f"[*] Получаю планировки у #{buildingId}..")
    url = f"https://www.workle.ru/api/v1/realty/plans/?buildingId={buildingId}&skip=0&take=12&spid={spid}"
    r = requests.get(url=url, headers=headers)
    print(f"[*] Ответ на запрос: {r.status_code}.")
    jString = json.loads(r.text)
    data = list()
    print(f"[/] Стуктурирую информацию..")
    for h in range(len(jString['list'])):
        current_offer = {
            "image": jString['list'][h]['image'],
            "floor": jString['list'][h]['floor'],
            "area": jString['list'][h]['area'],
            "rooms": jString['list'][h]['rooms'],
            "price": jString['list'][h]['price'],
        }
        data.append(current_offer)
    print(f"[$] Планировки структурированны успешно!")
    return data


def writeExcelData(data, filename="output"):
    print("[*] Начинаю создавать Excel файл..")
    # Создание нового файла Excel
    workbook = openpyxl.Workbook()

    # Создание листа "SPB"
    worksheet_spb = workbook.create_sheet(title="Санкт-Петербург")

    # Создание листа "MOSCOW"
    worksheet_moscow = workbook.create_sheet(title="Москва")

    # Запись заголовков столбцов
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        worksheet_spb.cell(row=1, column=col_num).value = header

    print("[/] Начинаю запись строк из полученных данных..")
    current_row = 2
    for h in range(len(data)):
        worksheet_spb[f'A{current_row}'].value = data[h]['BuildingId']
        worksheet_spb[f'B{current_row}'].value = data[h]['MainImages']
        worksheet_spb[f'C{current_row}'].value = data[h]['House']
        worksheet_spb[f'D{current_row}'].value = data[h]['MinPrice']
        worksheet_spb[f'E{current_row}'].value = data[h]['MaxPrice']
        worksheet_spb[f'F{current_row}'].value = data[h]['Metro']
        worksheet_spb[f'G{current_row}'].value = data[h]['MetroTime']
        worksheet_spb[f'H{current_row}'].value = data[h]['Region']
        worksheet_spb[f'I{current_row}'].value = data[h]['Zastroyshik']
        worksheet_spb[f'J{current_row}'].value = data[h]['Sdacha']
        worksheet_spb[f'K{current_row}'].value = data[h]['FloorsTotal']
        worksheet_spb[f'L{current_row}'].value = data[h]['Otdelka']
        worksheet_spb[f'M{current_row}'].value = data[h]['HouseType']
        worksheet_spb[f'N{current_row}'].value = data[h]['Description']
        worksheet_spb[f'O{current_row}'].value = data[h]['MinArea']
        worksheet_spb[f'P{current_row}'].value = data[h]['MaxArea']
        worksheet_spb[f'Q{current_row}'].value = data[h]['SubBuildings']
        current_row += 1

    # Сохранение файла
    workbook.save(f"{filename}.xlsx")
    print("[$] Excel файл успешно создан!")


def main():
    HouseIDs = getFirstHouseIds(78, 266, 5444)
    data = list()
    for h in range(len(HouseIDs)):
        data.append(getCurrentHouseInfo(HouseIDs[h]))

    writeExcelData(data, "test2")
    # writeExcelData(data=getCurrentHouseInfo(13890), filename="testOut")

    # getFirstHouseIds(78, 266, 5444)
    # getCurrentHouseInfo(13890)
    # print(getSubBuildings(13890))
    # dlImage("http://img.nmarket.pro/photo/pid/C9BA03B8-9D6D-452A-A2EA-CCA0A94E0BA4/?v=1&cid=&pid=C9BA03B8-9D6D-452A-A2EA-CCA0A94E0BA4&wpsid=52&hash=42b93f3fd690b5cc2cc97cd427a93463",
    #         "0",
    #         "13890",
    #         "main")


if __name__ == "__main__":
    main()
